from pathlib import Path
import logging
import json
import traceback
from typing import Optional, List, Any

Logger = logging.getLogger(__name__)

try:
    import pandas as pd
except Exception:
    pd = None

try:
    from openpyxl import load_workbook
except Exception:
    load_workbook = None  # type: ignore

try:
    from .extract_figures import extract_from_xlsx  # type: ignore
except Exception:
    extract_from_xlsx = None  # type: ignore


def _detect_header_row(df_raw, min_non_null: int) -> Optional[int]:
    for i, row in df_raw.iterrows():
        non_null = row.notna().sum()
        if non_null >= min_non_null:
            return int(i)
    return None


def _extract_table_bounds(df_raw):
    mask = df_raw.notna()
    if mask.values.sum() == 0:
        return None
    rows = mask.any(axis=1)
    cols = mask.any(axis=0)
    top = int(rows.idxmax())
    bottom = int(rows[::-1].idxmax())
    left = int(cols.idxmax())
    right = int(cols[::-1].idxmax())
    return top, bottom, left, right


def convert(
    xlsx_path: str | Path,
    json_path: str | Path,
    *,
    sheet_name: Optional[str | int] = 0,
    try_all_sheets: bool = False,
    ndjson: bool = False,
    orient: str = "records",
    min_non_null: int = 2,
    detect_tables: bool = True,
    skip_top_rows: int = 0,
    flatten: bool = False,
    encoding: str = "utf-8",
    on_error: str = "raise",
    error_log: Optional[str] = None,
) -> Optional[List[Path]]:
    """Convertit un XLSX en JSON/NDJSON.

    - `try_all_sheets`: écrit un fichier par feuille si True.
    - `ndjson`: si True, écrit en NDJSON (une ligne JSON par enregistrement).
    - `flatten`: utilise `pandas.json_normalize` pour aplatir les objets imbriqués.
    Retourne la liste des fichiers écrits, ou None en cas d'erreur si `on_error!='raise'`.
    """

    xlsx_path = Path(xlsx_path)
    json_path = Path(json_path)

    if not xlsx_path.exists():
        msg = f"Fichier XLSX introuvable: {xlsx_path}"
        if on_error == "raise":
            raise FileNotFoundError(xlsx_path)
        Logger.warning(msg)
        if error_log:
            Path(error_log).write_text(msg)
        return None

    results: List[Path] = []

    figures_dir = json_path.parent / f"{json_path.stem}_figures"
    manifest = None
    try:
        if extract_from_xlsx is not None:
            manifest = extract_from_xlsx(xlsx_path, figures_dir)
    except Exception:
        Logger.exception("Failed to extract figures from XLSX")

    try:
        if pd is not None:
            if try_all_sheets:
                xls = pd.read_excel(xlsx_path, sheet_name=None, header=None, engine="openpyxl")
                for sname, df_raw in xls.items():
                    if skip_top_rows:
                        df_raw = df_raw.iloc[skip_top_rows:]
                    header_row = _detect_header_row(df_raw, min_non_null)
                    if header_row is not None:
                        df = pd.read_excel(xlsx_path, sheet_name=sname, header=header_row, engine="openpyxl")
                    else:
                        df = df_raw.dropna(axis=0, how="all").dropna(axis=1, how="all")
                        df.columns = [f"col{i+1}" for i in range(df.shape[1])]

                    if flatten:
                        df = pd.json_normalize(df.to_dict(orient="records"))

                    out = json_path.with_name(f"{json_path.stem}_{str(sname)[:20]}.json")
                    if ndjson:
                        with open(out, "w", encoding=encoding) as f:
                            for rec in df.where(pd.notnull(df), None).to_dict(orient="records"):
                                f.write(json.dumps(rec, ensure_ascii=False) + "\n")
                    else:
                        with open(out, "w", encoding=encoding) as f:
                            payload = df.where(pd.notnull(df), None).to_dict(orient=orient)
                            # include manifest reference
                            if manifest is not None:
                                out_man = figures_dir / "manifest.json"
                                payload = {"_figures_manifest": str(out_man), "data": payload}
                            json.dump(payload, f, ensure_ascii=False, indent=2)
                    results.append(out)
                return results
            else:
                df_raw = pd.read_excel(xlsx_path, sheet_name=sheet_name, header=None, engine="openpyxl")
                if skip_top_rows:
                    df_raw = df_raw.iloc[skip_top_rows:]

                header_row = _detect_header_row(df_raw, min_non_null)
                if detect_tables:
                    bounds = _extract_table_bounds(df_raw)
                    if bounds:
                        top, bottom, left, right = bounds
                        df_table = df_raw.iloc[top: bottom + 1, left: right + 1]
                    else:
                        df_table = df_raw
                else:
                    df_table = df_raw

                if header_row is not None:
                    df = pd.read_excel(xlsx_path, sheet_name=sheet_name, header=header_row, engine="openpyxl")
                else:
                    first_row = list(df_table.iloc[0]) if not df_table.empty else []
                    non_null = sum(1 for v in first_row if pd.notna(v))
                    if non_null >= min_non_null and any(isinstance(v, str) for v in first_row):
                        cols = [str(x).strip() or f"col{i+1}" for i, x in enumerate(first_row)]
                        df = df_table.copy()
                        df.columns = cols
                        df = df.iloc[1:]
                    else:
                        df = df_table.copy()
                        df.columns = [f"col{i+1}" for i in range(df.shape[1])]

                if flatten:
                    df = pd.json_normalize(df.where(pd.notnull(df), None).to_dict(orient="records"))

                out = json_path
                if ndjson:
                    with open(out, "w", encoding=encoding) as f:
                        for rec in df.where(pd.notnull(df), None).to_dict(orient="records"):
                            f.write(json.dumps(rec, ensure_ascii=False) + "\n")
                else:
                    with open(out, "w", encoding=encoding) as f:
                        payload = df.where(pd.notnull(df), None).to_dict(orient=orient)
                        if manifest is not None:
                            out_man = figures_dir / "manifest.json"
                            payload = {"_figures_manifest": str(out_man), "data": payload}
                        json.dump(payload, f, ensure_ascii=False, indent=2)
                results.append(out)
                return results
        else:
            # fallback to openpyxl
            if load_workbook is None:
                raise RuntimeError("Ni pandas ni openpyxl disponibles pour lire XLSX")
            wb = load_workbook(filename=str(xlsx_path), read_only=True, data_only=True)
            sheet_keys = [sheet_name] if not try_all_sheets else wb.sheetnames
            for s in sheet_keys:
                ws = wb[s]
                rows = list(ws.iter_rows(values_only=True))
                if skip_top_rows:
                    rows = rows[skip_top_rows:]
                header_row = None
                for i, r in enumerate(rows):
                    non_null = sum(1 for c in r if c is not None)
                    if non_null >= min_non_null:
                        header_row = i
                        break
                if header_row is None:
                    header = [f"col{i+1}" for i in range(len(rows[0]) if rows else 0)]
                    data_rows = rows
                else:
                    header = [str(x).strip() if x is not None else f"col{i+1}" for i, x in enumerate(rows[header_row])]
                    data_rows = rows[header_row + 1 :]

                recs = []
                for r in data_rows:
                    rec = {header[i]: (r[i] if r and i < len(r) else None) for i in range(len(header))}
                    recs.append(rec)

                out = json_path.with_name(f"{json_path.stem}_{str(s)[:20]}.json") if try_all_sheets else json_path
                if ndjson:
                    with open(out, "w", encoding=encoding) as f:
                        for rec in recs:
                            f.write(json.dumps(rec, ensure_ascii=False) + "\n")
                else:
                    with open(out, "w", encoding=encoding) as f:
                        json.dump(recs, f, ensure_ascii=False, indent=2)
                results.append(out)
            return results
    except Exception as exc:
        msg = f"Erreur conversion XLSX→JSON: {exc}\n{traceback.format_exc()}"
        if on_error == "raise":
            raise
        Logger.exception(msg)
        if error_log:
            try:
                Path(error_log).write_text(msg)
            except Exception:
                Logger.warning("Impossible d'écrire error_log: %s", error_log)
        return None


def xlsx_to_json(*a, **k):
    return convert(*a, **k)


def xlsx_to_json_multisheet(*a, **k):
    return convert(*a, try_all_sheets=True, **k)


def xlsx_to_json_safe(*a, **k):
    return convert(*a, **k)


__all__ = ["convert", "xlsx_to_json", "xlsx_to_json_multisheet", "xlsx_to_json_safe"]
