from __future__ import annotations

from pathlib import Path
from typing import Any, Dict, List, Optional
import json
import logging
import traceback

try:
    import pandas as pd  # optionnel
except Exception:
    pd = None

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.utils import get_column_letter
except Exception:
    Workbook = None  # type: ignore
    load_workbook = None  # type: ignore
    get_column_letter = None  # type: ignore

Logger = logging.getLogger(__name__)


# =========================
# Exceptions
# =========================

class JSON2XLSXError(RuntimeError):
    pass


# =========================
# Utils
# =========================

def _flatten_dict(d: Dict[str, Any], sep: str) -> Dict[str, Any]:
    """
    Aplatissement déterministe :
    - dict → clés composées
    - list → JSON string
    """
    out: Dict[str, Any] = {}

    def _rec(x: Any, prefix: str) -> None:
        if isinstance(x, dict):
            for k, v in x.items():
                _rec(v, f"{prefix}{sep}{k}" if prefix else k)
        elif isinstance(x, list):
            out[prefix] = json.dumps(x, ensure_ascii=False)
        else:
            out[prefix] = x

    _rec(d, "")
    return out


def _autosize_worksheet(ws) -> None:
    if get_column_letter is None:
        return

    for i, col in enumerate(ws[1], start=1):
        max_len = 0
        for cell in ws[get_column_letter(i)]:
            v = cell.value
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        ws.column_dimensions[get_column_letter(i)].width = min(max_len + 2, 100)


def _write_openpyxl_single(
    rows: List[Dict[str, Any]],
    path: Path,
    *,
    sheet_name: str,
    autosize: bool,
) -> None:
    if Workbook is None:
        raise JSON2XLSXError("openpyxl requis pour écrire un XLSX")

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]

    headers = sorted({k for r in rows for k in r.keys()})
    ws.append(headers)

    for r in rows:
        ws.append([r.get(h) for h in headers])

    if autosize:
        _autosize_worksheet(ws)

    wb.save(path)


# =========================
# API principale
# =========================

def convert(
    json_input: str | Path,
    xlsx_output: str | Path,
    *,
    encoding: str = "utf-8",
    flatten: bool = False,
    flat_sep: str = ".",
    multisheet: bool = False,
    autosize: bool = True,
    on_error: str = "raise",  # "raise" | "warn" | "ignore"
    error_log: Optional[str | Path] = None,
) -> bool:
    """
    Convertit un JSON en XLSX.

    Cas supportés :
    - JSON liste de dicts → 1 feuille
    - JSON dict → 1 feuille
    - multisheet=True + dict[str, list[dict]] → plusieurs feuilles

    pandas utilisé si disponible, sinon fallback openpyxl.
    """

    json_path = Path(json_input)
    xlsx_path = Path(xlsx_output)

    if not json_path.exists():
        msg = f"Fichier JSON introuvable: {json_path}"
        if on_error == "raise":
            raise FileNotFoundError(json_path)
        Logger.warning(msg)
        return False

    try:
        with open(json_path, "r", encoding=encoding) as f:
            data = json.load(f)
    except Exception as exc:
        msg = f"Erreur lecture JSON: {exc}"
        Logger.exception(msg)
        if error_log:
            Path(error_log).write_text(msg + "\n" + traceback.format_exc())
        if on_error == "raise":
            raise
        return False

    xlsx_path.parent.mkdir(parents=True, exist_ok=True)

    # =========================
    # MULTI-SHEET
    # =========================

    if multisheet and isinstance(data, dict):
        try:
            if pd is not None:
                with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
                    for name, rows in data.items():
                        if not isinstance(rows, list):
                            continue
                        if flatten:
                            df = pd.json_normalize(rows, sep=flat_sep)
                        else:
                            df = pd.DataFrame(rows)
                        df.to_excel(writer, sheet_name=str(name)[:31], index=False)

                if autosize and load_workbook is not None:
                    wb = load_workbook(xlsx_path)
                    for ws in wb.worksheets:
                        _autosize_worksheet(ws)
                    wb.save(xlsx_path)

            else:
                if Workbook is None:
                    raise JSON2XLSXError("openpyxl requis pour multisheet XLSX")

                wb = Workbook()
                wb.remove(wb.active)

                for name, rows in data.items():
                    if not isinstance(rows, list):
                        continue
                    rows_proc = []
                    for r in rows:
                        if isinstance(r, dict):
                            rows_proc.append(
                                _flatten_dict(r, flat_sep) if flatten else r
                            )
                        else:
                            rows_proc.append({"value": r})

                    ws = wb.create_sheet(title=str(name)[:31])
                    headers = sorted({k for r in rows_proc for k in r.keys()})
                    ws.append(headers)
                    for r in rows_proc:
                        ws.append([r.get(h) for h in headers])
                    if autosize:
                        _autosize_worksheet(ws)

                wb.save(xlsx_path)

            return True

        except Exception as exc:
            msg = f"Erreur écriture XLSX multisheet: {exc}"
            Logger.exception(msg)
            if error_log:
                Path(error_log).write_text(msg + "\n" + traceback.format_exc())
            if on_error == "raise":
                raise
            return False

    # =========================
    # SINGLE SHEET
    # =========================

    if isinstance(data, list):
        rows = data
    elif isinstance(data, dict):
        rows = [data]
    else:
        rows = [{"value": data}]

    rows_proc: List[Dict[str, Any]] = []
    for r in rows:
        if isinstance(r, dict):
            rows_proc.append(_flatten_dict(r, flat_sep) if flatten else r)
        else:
            rows_proc.append({"value": r})

    try:
        if pd is not None:
            df = pd.DataFrame(rows_proc)
            df.to_excel(xlsx_path, index=False)

            if autosize and load_workbook is not None:
                wb = load_workbook(xlsx_path)
                for ws in wb.worksheets:
                    _autosize_worksheet(ws)
                wb.save(xlsx_path)
            return True

    except Exception:
        Logger.debug("Pandas XLSX failed, fallback openpyxl")

    try:
        _write_openpyxl_single(
            rows_proc,
            xlsx_path,
            sheet_name="Sheet1",
            autosize=autosize,
        )
        return True

    except Exception as exc:
        msg = f"Impossible d'écrire XLSX: {exc}"
        Logger.exception(msg)
        if error_log:
            Path(error_log).write_text(msg + "\n" + traceback.format_exc())
        if on_error == "raise":
            raise
        return False


# =========================
# Alias
# =========================

def json_to_xlsx(*a, **k):
    return convert(*a, **k)


__all__ = ["convert", "json_to_xlsx", "JSON2XLSXError"]
