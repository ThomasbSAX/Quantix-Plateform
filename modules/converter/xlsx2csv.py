from pathlib import Path
import logging
import traceback
from typing import Optional, List, Tuple, Any

Logger = logging.getLogger(__name__)

try:
    import pandas as pd
except Exception:
    pd = None

try:
    from openpyxl import load_workbook
except Exception:
    load_workbook = None  # type: ignore

try:
    from .extract_figures import extract_from_xlsx  # type: ignore
except Exception:
    extract_from_xlsx = None  # type: ignore


def _detect_header_row(df_raw, min_non_null: int) -> Optional[int]:
    # prefer the first row with enough non-null values
    for i, row in df_raw.iterrows():
        non_null = row.notna().sum()
        if non_null >= min_non_null:
            return int(i)
    return None


def _extract_table_bounds(df_raw) -> Optional[Tuple[int, int, int, int]]:
    # find bounding box of non-empty cells
    mask = df_raw.notna()
    if mask.values.sum() == 0:
        return None
    rows = mask.any(axis=1)
    cols = mask.any(axis=0)
    top = int(rows.idxmax())
    bottom = int(rows[::-1].idxmax())
    left = int(cols.idxmax())
    right = int(cols[::-1].idxmax())
    return top, bottom, left, right


def _is_noise_row(row_values) -> bool:
    # heuristics: long gibberish, mostly non-alnum, or a single long token without spaces
    s = " ".join("" if v is None else str(v) for v in row_values).strip()
    if not s:
        return False
    # too short to be noise
    if len(s) < 6:
        return False
    # mostly non-alphanumeric
    alpha = sum(1 for c in s if c.isalnum())
    ratio = alpha / max(1, len(s))
    if ratio < 0.3 and len(s) > 10:
        return True
    # uppercase gibberish token
    tokens = s.split()
    for t in tokens:
        if len(t) >= 10 and t.isupper() and sum(1 for c in t if c.isalpha()) / len(t) > 0.8:
            return True
    return False


def _rows_from_worksheet(ws):
    # Return a list of rows (lists) with merged cell values filled in.
    rows = []
    for r in ws.iter_rows(values_only=True):
        rows.append([c for c in r])

    # fill merged cells using merged_cell_ranges
    try:
        merged = list(ws.merged_cells.ranges)
    except Exception:
        merged = []
    for rng in merged:
        try:
            min_row = rng.min_row - 1
            max_row = rng.max_row - 1
            min_col = rng.min_col - 1
            max_col = rng.max_col - 1
            if min_row < len(rows) and min_col < len(rows[min_row]):
                val = rows[min_row][min_col]
                for ri in range(min_row, min(max_row + 1, len(rows))):
                    for ci in range(min_col, min(max_col + 1, len(rows[ri]))):
                        rows[ri][ci] = val
        except Exception:
            continue
    return rows


def convert(
    xlsx_path: str | Path,
    csv_path: str | Path,
    *,
    sheet_name: Optional[str | int] = 0,
    try_all_sheets: bool = False,
    sep: str = ",",
    min_non_null: int = 2,
    detect_tables: bool = True,
    skip_top_rows: int = 0,
    encoding: str = "utf-8",
    on_error: str = "raise",
    error_log: Optional[str] = None,
) -> Optional[List[Path]]:
    """Convertit un XLSX en CSV en gérant de nombreux cas erronés.

    - `try_all_sheets`: si True, écrit un CSV par feuille (suffixe avec le nom).
    - `detect_tables`: tente d'isoler une table dans la feuille (ignore en-têtes/notes/graphes).
    - Retourne la liste des fichiers CSV écrits, ou None si échec et `on_error!='raise'`.
    """

    xlsx_path = Path(xlsx_path)
    csv_path = Path(csv_path)

    if not xlsx_path.exists():
        msg = f"Fichier XLSX introuvable: {xlsx_path}"
        if on_error == "raise":
            raise FileNotFoundError(xlsx_path)
        Logger.warning(msg)
        if error_log:
            Path(error_log).write_text(msg)
        return None

    results: List[Path] = []

    # prepare figures extraction folder
    figures_dir = csv_path.parent / f"{csv_path.stem}_figures"
    manifest = None
    try:
        if extract_from_xlsx is not None:
            manifest = extract_from_xlsx(xlsx_path, figures_dir)
    except Exception:
        # log and continue
        Logger.exception("Failed to extract figures from XLSX")

    def _write_df_to_csv(df, out_path: Path):
        try:
            df.to_csv(out_path, sep=sep, index=False, encoding=encoding)
            results.append(out_path)
        except Exception:
            # fallback: write via csv module
            import csv as _csv

            with open(out_path, "w", newline="", encoding=encoding) as fout:
                writer = _csv.writer(fout, delimiter=sep)
                # header
                writer.writerow(list(df.columns))
                for row in df.itertuples(index=False, name=None):
                    writer.writerow(["" if v is None else v for v in row])
            results.append(out_path)

    try:
        if pd is not None:
            # use pandas path
            if try_all_sheets:
                xls = pd.read_excel(xlsx_path, sheet_name=None, header=None, engine="openpyxl")
                for sname, df_raw in xls.items():
                    # skip initial rows
                    if skip_top_rows:
                        df_raw = df_raw.iloc[skip_top_rows:]
                    header_row = _detect_header_row(df_raw, min_non_null)
                    if header_row is None:
                        df = df_raw.dropna(axis=0, how="all").dropna(axis=1, how="all")
                        # ensure string columns
                        df.columns = [f"col{i+1}" for i in range(df.shape[1])]
                    else:
                        df = pd.read_excel(xlsx_path, sheet_name=sname, header=header_row, engine="openpyxl")
                        df = df.dropna(axis=0, how="all").dropna(axis=1, how="all")

                    out = csv_path.with_name(f"{csv_path.stem}_{str(sname)[:20]}.csv")
                    # if manifest exists, add manifest path column
                    if manifest is not None:
                        try:
                            manifest_path = figures_dir / "manifest.json"
                            df["__figures_manifest"] = str(manifest_path)
                        except Exception:
                            pass
                    _write_df_to_csv(df, out)
                return results
            else:
                df_raw = pd.read_excel(xlsx_path, sheet_name=sheet_name, header=None, engine="openpyxl")
                if skip_top_rows:
                    df_raw = df_raw.iloc[skip_top_rows:]

                # Prefer openpyxl-backed extraction for reliable merged-cell handling and shapes
                df_table = None
                if detect_tables and load_workbook is not None:
                    try:
                        wb = load_workbook(filename=str(xlsx_path), data_only=True)
                        ws = None
                        if isinstance(sheet_name, int):
                            ws = wb.worksheets[sheet_name]
                        else:
                            ws = wb[sheet_name]
                        rows = _rows_from_worksheet(ws)
                        import pandas as _pd

                        df_raw2 = _pd.DataFrame(rows)
                        if skip_top_rows:
                            df_raw2 = df_raw2.iloc[skip_top_rows:]
                        # remove noise rows (text boxes, footers etc.)
                        df_raw2 = df_raw2.dropna(axis=1, how="all")
                        clean_rows = [r for r in df_raw2.values.tolist() if not _is_noise_row(r)]
                        if clean_rows:
                            df_raw2 = _pd.DataFrame(clean_rows)
                        bounds = _extract_table_bounds(df_raw2)
                        if bounds:
                            top, bottom, left, right = bounds
                            df_table = df_raw2.iloc[top: bottom + 1, left: right + 1]
                        else:
                            df_table = df_raw2
                    except Exception:
                        df_table = df_raw
                else:
                    header_row = _detect_header_row(df_raw, min_non_null)
                    if detect_tables:
                        bounds = _extract_table_bounds(df_raw)
                        if bounds:
                            top, bottom, left, right = bounds
                            df_table = df_raw.iloc[top: bottom + 1, left: right + 1]
                        else:
                            df_table = df_raw
                    else:
                        df_table = df_raw

                # attempt to find header row within df_table
                header_row = _detect_header_row(df_table, min_non_null)
                if header_row is not None:
                    try:
                        df = pd.read_excel(xlsx_path, sheet_name=sheet_name, header=header_row + (0 if skip_top_rows == 0 else skip_top_rows), engine="openpyxl")
                        df = df.dropna(axis=0, how="all").dropna(axis=1, how="all")
                    except Exception:
                        # fallback: build df from df_table and use header row
                        cols = [str(x).strip() or f"col{i+1}" for i, x in enumerate(df_table.iloc[header_row])]
                        df = df_table.copy()
                        df.columns = cols
                        df = df.iloc[header_row + 1 :]
                        df = df.dropna(axis=0, how="all").dropna(axis=1, how="all")
                else:
                    # no header detected -> use first row of df_table as header if likely, else generate
                    first_row = list(df_table.iloc[0]) if not df_table.empty else []
                    non_null = sum(1 for v in first_row if pd.notna(v))
                    if non_null >= min_non_null and any(isinstance(v, str) for v in first_row):
                        cols = [str(x).strip() or f"col{i+1}" for i, x in enumerate(first_row)]
                        df = df_table.copy()
                        df.columns = cols
                        df = df.iloc[1:]
                    else:
                        df = df_table.copy()
                        df.columns = [f"col{i+1}" for i in range(df.shape[1])]
                out = csv_path
                # add manifest column to main sheet CSV if available
                if manifest is not None:
                    try:
                        manifest_path = figures_dir / "manifest.json"
                        df["__figures_manifest"] = str(manifest_path)
                    except Exception:
                        pass
                _write_df_to_csv(df, out)
                return results
        else:
            # fallback to openpyxl
            if load_workbook is None:
                raise RuntimeError("Ni pandas ni openpyxl disponibles pour lire XLSX")
            wb = load_workbook(filename=str(xlsx_path), read_only=True, data_only=True)
            sheet_keys = [sheet_name] if not try_all_sheets else wb.sheetnames
            for s in sheet_keys:
                ws = wb[s]
                rows = list(ws.iter_rows(values_only=True))
                if skip_top_rows:
                    rows = rows[skip_top_rows:]
                # detect header
                header_row = None
                for i, r in enumerate(rows):
                    non_null = sum(1 for c in r if c is not None)
                    if non_null >= min_non_null:
                        header_row = i
                        break
                if header_row is None:
                    header = [f"col{i+1}" for i in range(len(rows[0]) if rows else 0)]
                    data_rows = rows
                else:
                    header = [str(x).strip() if x is not None else f"col{i+1}" for i, x in enumerate(rows[header_row])]
                    data_rows = rows[header_row + 1 :]

                # write CSV
                import csv as _csv

                out = csv_path.with_name(f"{csv_path.stem}_{str(s)[:20]}.csv") if try_all_sheets else csv_path
                with open(out, "w", newline="", encoding=encoding) as fout:
                    writer = _csv.writer(fout, delimiter=sep)
                    writer.writerow(header)
                    for r in data_rows:
                        vals = ["" if v is None else v for v in r]
                        writer.writerow(vals)
                results.append(out)
            return results
    except Exception as exc:
        msg = f"Erreur conversion XLSX→CSV: {exc}\n{traceback.format_exc()}"
        if on_error == "raise":
            raise
        Logger.exception(msg)
        if error_log:
            try:
                Path(error_log).write_text(msg)
            except Exception:
                Logger.warning("Impossible d'écrire error_log: %s", error_log)
        return None


def xlsx_to_csv_clean(*a, **k):
    return convert(*a, **k)


__all__ = ["convert", "xlsx_to_csv_clean"]
