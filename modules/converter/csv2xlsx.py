from __future__ import annotations

import csv
import logging
from pathlib import Path
from typing import Any, Dict, List, Optional, Union

Logger = logging.getLogger(__name__)


# =========================
# Dépendance
# =========================

def _ensure_openpyxl() -> None:
    try:
        import openpyxl  # type: ignore
    except Exception as exc:
        raise ImportError(
            "openpyxl requis. Installez-le via `pip install openpyxl`"
        ) from exc


# =========================
# CSV utils (séparateurs)
# =========================

def _sniff_dialect(sample: str) -> csv.Dialect:
    """
    Détecte delimiter + quotechar + escapechar.
    Fallback explicite sur CSV standard.
    """
    try:
        return csv.Sniffer().sniff(sample)
    except Exception:
        class Fallback(csv.Dialect):
            delimiter = ","
            quotechar = '"'
            escapechar = None
            doublequote = True
            skipinitialspace = False
            lineterminator = "\n"
            quoting = csv.QUOTE_MINIMAL
        return Fallback()


def _infer_value(s: Optional[str]) -> Any:
    if s is None:
        return None
    v = s.strip()
    if v == "":
        return None

    low = v.lower()
    if low in {"null", "none", "nan"}:
        return None
    if low == "true":
        return True
    if low == "false":
        return False

    if v.isdigit() or (v.startswith("-") and v[1:].isdigit()):
        try:
            return int(v)
        except Exception:
            pass
    try:
        return float(v)
    except Exception:
        return v


def _autosize_columns(ws) -> None:
    dims: Dict[int, int] = {}
    for row in ws.iter_rows():
        for idx, cell in enumerate(row, 1):
            val = cell.value
            ln = len(str(val)) if val is not None else 0
            dims[idx] = max(dims.get(idx, 0), ln)

    for idx, width in dims.items():
        col = ws.cell(row=1, column=idx).column_letter
        ws.column_dimensions[col].width = min(max(width + 2, 8), 60)


# =========================
# Écriture d’une feuille
# =========================

def _write_single(
    wb,
    sheet_name: str,
    csv_path: Path,
    *,
    encoding: str,
    delimiter: Optional[str],
    has_header: Optional[bool],
    infer_types: bool,
    autosize: bool,
    startrow: int,
) -> None:
    if not csv_path.exists():
        raise FileNotFoundError(csv_path)

    try:
        try:
            fin = open(csv_path, "r", encoding=encoding, newline="")
        except UnicodeDecodeError:
            fin = open(csv_path, "r", encoding="latin-1", newline="")
    except Exception as exc:
        raise RuntimeError(f"Impossible d’ouvrir {csv_path}: {exc}") from exc

    with fin:
        sample = fin.read(8192)
        fin.seek(0)

        dialect = _sniff_dialect(sample)
        if delimiter is not None:
            dialect.delimiter = delimiter  # override explicite

        reader = csv.reader(
            fin,
            delimiter=dialect.delimiter,
            quotechar=dialect.quotechar,
            escapechar=dialect.escapechar,
            doublequote=dialect.doublequote,
            skipinitialspace=dialect.skipinitialspace,
        )

        try:
            first = next(reader)
        except StopIteration:
            wb.create_sheet(title=sheet_name)
            return

        if has_header is None:
            has_header = any(
                not c.replace(".", "").replace("-", "").isdigit()
                for c in first
            )

        if has_header:
            headers = [h.strip() or "field" for h in first]
        else:
            headers = [f"field{i+1}" for i in range(len(first))]
            fin.seek(0)
            reader = csv.reader(
                fin,
                delimiter=dialect.delimiter,
                quotechar=dialect.quotechar,
                escapechar=dialect.escapechar,
                doublequote=dialect.doublequote,
                skipinitialspace=dialect.skipinitialspace,
            )

        ws = wb.create_sheet(title=sheet_name)
        r = startrow

        for c, h in enumerate(headers, 1):
            ws.cell(row=r, column=c, value=h)
        r += 1

        for row in reader:
            if all((c or "").strip() == "" for c in row):
                continue

            if len(row) < len(headers):
                row += [""] * (len(headers) - len(row))
            elif len(row) > len(headers):
                row = row[:len(headers)]

            for c, v in enumerate(row, 1):
                ws.cell(
                    row=r,
                    column=c,
                    value=_infer_value(v) if infer_types else (v or None),
                )
            r += 1

        if autosize:
            _autosize_columns(ws)


# =========================
# API publique
# =========================

def convert(
    csv_path: Union[str, Path, List[Union[str, Path]], Dict[str, Union[str, Path]]],
    xlsx_path: Union[str, Path],
    *,
    encoding: str = "utf-8",
    delimiter: Optional[str] = None,
    has_header: Optional[bool] = None,
    infer_types: bool = True,
    autosize: bool = True,
    startrow: int = 1,
    overwrite: bool = True,
) -> None:
    """
    CSV → XLSX robuste.

    - détection fiable delimiter / quote / escape
    - override manuel possible
    - support multi-feuilles
    - typage automatique optionnel
    """

    _ensure_openpyxl()
    from openpyxl import Workbook

    xlsx_path = Path(xlsx_path)
    if xlsx_path.exists() and not overwrite:
        raise FileExistsError(xlsx_path)

    wb = Workbook()
    wb.remove(wb.active)

    sources: Dict[str, Path] = {}
    if isinstance(csv_path, dict):
        sources = {str(k): Path(v) for k, v in csv_path.items()}
    elif isinstance(csv_path, (list, tuple)):
        for i, p in enumerate(csv_path, 1):
            sources[f"Sheet{i}"] = Path(p)
    else:
        sources["Sheet1"] = Path(csv_path)

    for sheet, path in sources.items():
        _write_single(
            wb,
            sheet,
            path,
            encoding=encoding,
            delimiter=delimiter,
            has_header=has_header,
            infer_types=infer_types,
            autosize=autosize,
            startrow=startrow,
        )

    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(xlsx_path)


__all__ = ["convert"]
